import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import re
import traceback
import os
from datetime import datetime

class ExcelHandler:
    def __init__(self, file_path):
        self.file_path = file_path
        self.target_green_rgb = '92D050'
        self.logs = []
        try:
            self.wb = openpyxl.load_workbook(self.file_path, data_only=True)
            self.wb_formula = openpyxl.load_workbook(self.file_path, data_only=False)
            self._log(f"Loaded workbook: {file_path}")
        except Exception as e:
            self._log(f"Error loading workbook {file_path}: {e}")
            self.wb = None
            self.wb_formula = None
        self.route_options_cache = None

    def _log(self, msg):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"[{timestamp}] {msg}"
        self.logs.append(log_entry)
        print(log_entry)

    def get_route_options(self):
        if self.route_options_cache:
            return self.route_options_cache
        
        if not self.wb: return {}
            
        sheets = ['WAHL-Customer', 'VENDOR-WAHL', 'WAHL-DGWA']
        options = {}
        
        for sheet_name in sheets:
            if sheet_name not in self.wb.sheetnames: continue
            ws = self.wb[sheet_name]
            header_row, map_col, from_col, to_col = self._find_header_info(ws)
            if not map_col: continue
            
            for r in range(header_row + 1, ws.max_row + 1):
                node_val = ws.cell(r, map_col).value
                if node_val is None: continue
                
                node = str(node_val).strip()
                frm = str(ws.cell(r, from_col).value).strip() if from_col and ws.cell(r, from_col).value else ""
                to = str(ws.cell(r, to_col).value).strip() if to_col and ws.cell(r, to_col).value else ""
                
                if node:
                    if node not in options:
                        options[node] = {'locations': [], 'details': [], 'sheet': sheet_name}
                    
                    loc_str = f"{frm} -> {to}"
                    if loc_str not in options[node]['locations']:
                        options[node]['locations'].append(loc_str)
                        options[node]['details'].append({"from": frm, "to": to})
        
        for node in options:
            options[node]['locations'].sort()
        
        # Special handling for E node - ensure it has both directions
        if 'E' in options:
            existing_locs = set(options['E']['locations'])
            has_wahl_wadg = any('WAHL' in loc and 'WADG' in loc for loc in existing_locs)
            
            # If E node exists but missing one direction, add it
            if has_wahl_wadg:
                if 'WAHL -> WADG' not in existing_locs:
                    options['E']['locations'].append('WAHL -> WADG')
                    options['E']['details'].append({"from": "WAHL", "to": "WADG"})
                if 'WADG -> WAHL' not in existing_locs:
                    options['E']['locations'].append('WADG -> WAHL')
                    options['E']['details'].append({"from": "WADG", "to": "WAHL"})
                options['E']['locations'].sort()
            
        self.route_options_cache = options
        return options

    def get_node_fields(self, node, location_str):
        sheet_name = self._get_sheet_for_node(node)
        if not sheet_name: return []
        
        ws = self.wb[sheet_name]
        header_row, _, _, to_col = self._find_header_info(ws)
        summary_col = self._get_col_by_header(ws, header_row, 'SUMMARY')
        if not summary_col or not to_col: return []

        matching_rows = []
        loc_parts = [s.strip() for s in location_str.split('->')]
        if len(loc_parts) != 2: return []
        frm_target, to_target = loc_parts
        
        _, map_col_idx, from_col_idx, to_col_idx = self._find_header_info(ws)
        
        for r in range(header_row + 1, ws.max_row + 1):
            row_node = str(ws.cell(r, map_col_idx).value).strip() if ws.cell(r, map_col_idx).value else ""
            row_frm = str(ws.cell(r, from_col_idx).value).strip() if ws.cell(r, from_col_idx).value else ""
            row_to = str(ws.cell(r, to_col_idx).value).strip() if ws.cell(r, to_col_idx).value else ""
            
            if row_node == node and row_frm == frm_target and row_to == to_target:
                matching_rows.append(r)

        if not matching_rows: return []

        fields = []
        for c in range(to_col + 1, summary_col + 1):
            title = ws.cell(header_row, c).value
            if not title: continue
            
            display_title = "Shipping method" if title == "SUMMARY" else title
            unique_values = set()
            for r in matching_rows:
                val = ws.cell(r, c).value
                if val is not None and str(val).strip().upper() != "N/A" and str(val).strip() != "":
                    if title == "SUMMARY":
                        if val == 'A': val = 'Ocean'
                        elif val == 'B': val = 'Air'
                        elif val == 'C': val = 'Land'
                    unique_values.add(val)
            
            if unique_values:
                fields.append({
                    "name": title,
                    "display_name": display_title,
                    "options": sorted(list(unique_values))
                })
        
        return fields

    def calculate(self, selections):
        self.logs = []
        results = []
        total_cost = 0
        all_lt_strings = []
        
        self._log("=" * 60)
        self._log("CALCULATION START")
        self._log("=" * 60)
        
        try:
            for idx, sel in enumerate(selections):
                node = sel.get('node')
                location = sel.get('location')
                inputs = sel.get('inputs', {})
                
                self._log(f"\n--- SECTION {idx+1} ---")
                self._log(f"User Selection: Node={node}, Location={location}")
                self._log(f"User Inputs: {inputs}")
                
                if not node or not location: 
                    self._log("ERROR: Missing node or location")
                    continue
                
                sheet_name = self._get_sheet_for_node(node)
                if not sheet_name: 
                    self._log(f"ERROR: No sheet found for node {node}")
                    continue
                
                self._log(f"Using sheet: {sheet_name}")
                    
                ws = self.wb[sheet_name]
                ws_formula = self.wb_formula[sheet_name]
                header_row, map_col, from_col, to_col = self._find_header_info(ws)
                self._log(f"Header row: {header_row}, MAP col: {map_col}, From col: {from_col}, To col: {to_col}")
                
                frm_target, to_target = [s.strip() for s in location.split('->')]
                self._log(f"Looking for: From='{frm_target}', To='{to_target}'")
                
                target_row = None
                # Try exact match first
                for r in range(header_row + 1, ws.max_row + 1):
                    if self._row_matches_exact(ws, r, header_row, map_col, from_col, to_col, node, frm_target, to_target, inputs):
                        target_row = r
                        self._log(f"EXACT MATCH found at row {r}")
                        break
                
                if target_row:
                    cost, lt_str, breakdown, log_details = self._extract_data_from_row(ws, ws_formula, sheet_name, target_row, header_row, inputs)
                    results.append({"node": node, "cost": cost, "lt": lt_str, "breakdown": breakdown})
                    total_cost += cost
                    if lt_str:
                        all_lt_strings.append(str(lt_str))
                    self._log(f"Extracted: Cost={cost}, LT='{lt_str}'")
                else:
                    # Try partial match (ignore PALLET QTY, CBM, G/W)
                    self._log("No exact match, trying partial match (ignoring PALLET QTY, CBM, G/W)...")
                    for r in range(header_row + 1, ws.max_row + 1):
                        match_result = self._row_matches_partial(ws, r, header_row, map_col, from_col, to_col, node, frm_target, to_target, inputs)
                        if match_result:
                            target_row = r
                            self._log(f"PARTIAL MATCH found at row {r}")
                            break
                    
                    if target_row:
                        # Calculate cost using formula with user inputs
                        cost, lt_str, breakdown, log_details = self._calculate_with_formula(ws, ws_formula, sheet_name, target_row, header_row, inputs)
                        results.append({"node": node, "cost": cost, "lt": lt_str, "breakdown": breakdown})
                        total_cost += cost
                        if lt_str:
                            all_lt_strings.append(str(lt_str))
                        self._log(f"Calculated (formula): Cost={cost}, LT='{lt_str}'")
                    else:
                        self._log(f"ERROR: No match found")
                        results.append({"node": node, "cost": 0, "lt": "", "breakdown": None, "error": f"未找到匹配: {frm_target} -> {to_target}"})
                        
        except Exception as e:
            self._log(f"EXCEPTION: {str(e)}")
            self._log(traceback.format_exc())

        total_lt = self._aggregate_lt(all_lt_strings)
        
        self._log("\n" + "=" * 60)
        self._log(f"TOTAL COST: {total_cost}")
        self._log(f"TOTAL LT: {total_lt}")
        self._log(f"LT Components: {all_lt_strings}")
        self._log("=" * 60)

        return {
            "node_results": results,
            "total_cost": total_cost,
            "total_lt": total_lt,
            "logs": self.logs
        }

    def _is_date_format(self, val):
        """Check if value looks like a date/days format"""
        if not val:
            return False
        s = str(val).lower().strip()
        has_numbers = bool(re.search(r'\d', s))
        has_day = 'day' in s or 'days' in s
        return has_numbers and has_day

    def _aggregate_lt(self, lt_strings):
        total_min = 0
        total_max = 0
        
        for lt in lt_strings:
            if not lt: continue
            s = str(lt).strip()
            self._log(f"  Parsing LT: '{s}'")
            nums = re.findall(r'\d+', s)
            self._log(f"    Found numbers: {nums}")
            if len(nums) == 1:
                val = int(nums[0])
                total_min += val
                total_max += val
            elif len(nums) >= 2:
                total_min += int(nums[0])
                total_max += int(nums[1])
        
        if total_min == 0 and total_max == 0:
            return "N/A"
        elif total_min == total_max:
            return f"{total_min} Days"
        else:
            return f"{total_min}-{total_max} Days"

    def _find_header_info(self, ws):
        for r in range(1, 6):
            row_values = []
            for c in range(1, min(ws.max_column + 1, 30)):
                row_values.append(ws.cell(r, c).value)
            
            if 'MAP' in row_values:
                map_col = from_col = to_col = None
                for idx, val in enumerate(row_values):
                    if val == 'MAP': map_col = idx + 1
                    elif val == 'From': from_col = idx + 1
                    elif val == 'To': to_col = idx + 1
                return r, map_col, from_col, to_col
        return 2, None, None, None

    def _get_sheet_for_node(self, node):
        options = self.get_route_options()
        return options.get(node, {}).get('sheet')

    def _row_matches_exact(self, ws, r, header_row, map_col, from_col, to_col, node, frm, to, inputs):
        """Check if row matches exactly including all input fields."""
        node_cell = ws.cell(r, map_col).value
        if not node_cell: return False
        if str(node_cell).strip() != node: return False
        
        frm_cell = ws.cell(r, from_col).value
        if not frm_cell or str(frm_cell).strip() != frm: return False

        to_cell = ws.cell(r, to_col).value
        if not to_cell or str(to_cell).strip() != to: return False
        
        for field, val in inputs.items():
            if not val or str(val).lower() == 'n/a': continue
            
            col = self._get_col_by_header(ws, header_row, field)
            if col:
                row_val = ws.cell(r, col).value
                if row_val is None:
                    row_val = ws.cell(r+1, col).value

                input_val = val
                if field == 'SUMMARY':
                    input_val = 'A' if val == 'Ocean' else ('B' if val == 'Air' else 'C')
                
                if str(row_val).strip() != str(input_val).strip():
                    return False
        return True

    def _row_matches_partial(self, ws, r, header_row, map_col, from_col, to_col, node, frm, to, inputs):
        """Check if row matches, ignoring PALLET QTY, CBM, G/W fields.
        Stricter logic: if user didn't input a field but Excel has a value, fail.
        """
        node_cell = ws.cell(r, map_col).value
        if not node_cell: return False
        if str(node_cell).strip() != node: return False
        
        frm_cell = ws.cell(r, from_col).value
        if not frm_cell or str(frm_cell).strip() != frm: return False

        to_cell = ws.cell(r, to_col).value
        if not to_cell or str(to_cell).strip() != to: return False
        
        # Skip these fields for partial matching
        skip_fields = ['PALLET QTY', 'CBM', 'G/W', 'GW']
        
        self._log(f"  检查第 {r} 行的部分匹配")
        
        # Get all field columns from Excel (between 'To' and 'SUMMARY')
        summary_col = self._get_col_by_header(ws, header_row, 'SUMMARY')
        to_col_idx = self._get_col_by_header(ws, header_row, 'To')
        
        if not summary_col or not to_col_idx:
            self._log(f"    错误：找不到 SUMMARY 或 To 列")
            return False
        
        # Iterate through each field column
        for c in range(to_col_idx + 1, summary_col + 1):
            field_header = ws.cell(header_row, c).value
            if not field_header:
                continue
            
            # Skip the special fields
            if field_header in skip_fields:
                self._log(f"    字段 '{field_header}': 在跳过列表中，跳过")
                continue
            
            # Get Excel value from row (check both merged rows)
            excel_val = ws.cell(r, c).value
            if excel_val is None:
                excel_val = ws.cell(r + 1, c).value
            
            # Get user input value
            user_val = inputs.get(field_header)
            
            # Handle SUMMARY field special conversion
            if field_header == 'SUMMARY' and user_val:
                user_val_converted = 'A' if user_val == 'Ocean' else ('B' if user_val == 'Air' else 'C')
            else:
                user_val_converted = user_val
            
            # Check logic:
            # 1. User didn't input (empty/None)
            if not user_val or str(user_val).strip() == '':
                # Excel has a value (not empty/N/A) → FAIL
                if excel_val and str(excel_val).strip() not in ['', 'N/A']:
                    self._log(f"    字段 '{field_header}': 用户未输入，Excel 为 '{excel_val}' - 失败")
                    return False
                else:
                    self._log(f"    字段 '{field_header}': 用户未输入，Excel 也为空 - 跳过")
            # 2. User did input
            else:
                # Excel is empty/N/A but user has input → FAIL (stricter matching)
                if not excel_val or str(excel_val).strip() in ['', 'N/A']:
                    self._log(f"    字段 '{field_header}': Excel 为空，但用户输入='{user_val}' - 失败")
                    return False
                
                # Normal match check
                if str(excel_val).strip() != str(user_val_converted).strip():
                    self._log(f"    字段 '{field_header}': 不匹配 - Excel='{excel_val}', 用户='{user_val}' - 失败")
                    return False
                else:
                    self._log(f"    字段 '{field_header}': 匹配 - Excel='{excel_val}', 用户='{user_val}' - 成功")
        
        self._log(f"  第 {r} 行部分匹配成功")
        return True

    def _get_col_by_header(self, ws, header_row, header_name):
        for c in range(1, ws.max_column + 1):
            if ws.cell(header_row, c).value == header_name:
                return c
        return None

    def _parse_min_value(self, cell_text, sheet_name, inputs):
        """
        Parse MIN logic from cell text.
        Returns calculated value based on MIN rules.
        """
        if not cell_text or 'MIN' not in str(cell_text).upper():
            return None
            
        text = str(cell_text).upper()
        self._log(f"  Parsing MIN in: '{cell_text}'")
        
        # Extract MIN value
        min_match = re.search(r'MIN\s*(\d+(?:\.\d+)?)', text)
        if not min_match:
            return None
        min_val = float(min_match.group(1))
        
        # Extract base rate (number before MIN)
        base_match = re.search(r'(\d+(?:\.\d+)?)', text)
        if not base_match:
            return None
        base_rate = float(base_match.group(1))
        
        pallet_qty = float(inputs.get('PALLET QTY', 0) or 0)
        cbm = float(inputs.get('CBM', 0) or 0)
        gw = float(inputs.get('G/W', 0) or inputs.get('GW', 0) or 0)
        
        calculated = 0
        
        if sheet_name == 'VENDOR-WAHL':
            if 'CBM' in text:
                calculated = base_rate * 7.8 * cbm
                self._log(f"    VENDOR-WAHL CBM: {base_rate} * 7.8 * {cbm} = {calculated}")
            elif 'KG' in text:
                calculated = base_rate * gw
                self._log(f"    VENDOR-WAHL KG: {base_rate} * {gw} = {calculated}")
            else:
                calculated = base_rate * pallet_qty
        else:  # WAHL-Customer
            calculated = base_rate * pallet_qty
            self._log(f"    WAHL-Customer: {base_rate} * {pallet_qty} = {calculated}")
        
        result = max(calculated, min_val)
        self._log(f"    Compare: calculated={calculated}, MIN={min_val}, result={result}")
        return result

    def _evaluate_cell_formula(self, ws, ws_formula, formula, row, header_row, inputs, sheet_name):
        """Evaluate a cell formula, replacing references to PALLET QTY, CBM, G/W with user inputs."""
        if not formula or not isinstance(formula, str):
            return 0
        
        try:
            # Get user input values
            pallet_qty = float(inputs.get('PALLET QTY', 0) or 0)
            cbm = float(inputs.get('CBM', 0) or 0)
            gw = float(inputs.get('G/W', 0) or inputs.get('GW', 0) or 0)
            
            # Get column indices for user input fields
            pallet_col = self._get_col_by_header(ws, header_row, 'PALLET QTY')
            cbm_col = self._get_col_by_header(ws, header_row, 'CBM')
            gw_col = self._get_col_by_header(ws, header_row, 'G/W')
            
            # Simple formula parsing for common patterns like =A1*B2, =A1+B2, =A1
            formula_str = formula[1:]  # Remove leading '='
            
            # Find all cell references (e.g., A1, AB12)
            cell_refs = re.findall(r'([A-Z]+)(\d+)', formula_str)
            
            result_str = formula_str
            
            for col_letter, row_num in cell_refs:
                col_idx = openpyxl.utils.column_index_from_string(col_letter)
                ref_row = int(row_num)
                
                # Check if this reference is to a user input column
                if col_idx == pallet_col:
                    val = pallet_qty
                    self._log(f"      替换 {col_letter}{row_num} (PALLET QTY) = {val}")
                elif col_idx == cbm_col:
                    val = cbm
                    self._log(f"      替换 {col_letter}{row_num} (CBM) = {val}")
                elif col_idx == gw_col:
                    val = gw
                    self._log(f"      替换 {col_letter}{row_num} (G/W) = {val}")
                else:
                    # Get value from Excel
                    val = ws.cell(ref_row, col_idx).value
                    if val is None:
                        val = 0
                    try:
                        val = float(val)
                    except (ValueError, TypeError):
                        val = 0
                
                # Replace cell reference with value
                result_str = result_str.replace(f"{col_letter}{row_num}", str(val), 1)
            
            # Evaluate the expression
            # Only allow safe characters
            safe_chars = set('0123456789.+-*/() ')
            if all(c in safe_chars for c in result_str):
                result = eval(result_str)
                return float(result)
            else:
                self._log(f"      不安全的公式表达式: {result_str}")
                return 0
                
        except Exception as e:
            self._log(f"      公式计算错误: {e}")
            return 0

    def _calculate_with_formula(self, ws, ws_formula, sheet_name, row, header_row, inputs):
        """Calculate E2E Cost using formula with user inputs for partial match."""
        e2e_cost_col = self._get_col_by_header(ws, header_row, 'E2E Cost')
        e2e_lt_col = self._get_col_by_header(ws, header_row, 'E2E Lead Time')
        
        self._log(f"Calculating with formula at row {row}")
        
        # Get the formula from the cell
        formula_cell = ws_formula.cell(row, e2e_cost_col)
        formula = formula_cell.value
        self._log(f"Formula: {formula}")
        
        total_cost = 0
        
        # If formula is SUM(range), parse and calculate each cell
        if formula and isinstance(formula, str) and formula.startswith('=SUM'):
            # Extract range like N30:AG30
            range_match = re.search(r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', formula)
            if range_match:
                start_col = range_match.group(1)
                start_row = int(range_match.group(2))
                end_col = range_match.group(3)
                end_row = int(range_match.group(4))
                
                start_col_idx = openpyxl.utils.column_index_from_string(start_col)
                end_col_idx = openpyxl.utils.column_index_from_string(end_col)
                
                self._log(f"SUM range: {start_col}{start_row} to {end_col}{end_row}")
                
                for c in range(start_col_idx, end_col_idx + 1):
                    # Get value from both rows of merged cell
                    row1_val = ws.cell(row, c).value
                    row2_val = ws.cell(row + 1, c).value
                    header_val = ws.cell(header_row, c).value
                    
                    # Check if row2 cell has a formula
                    formula_val = ws_formula.cell(row + 1, c).value
                    
                    cell_contribution = 0
                    
                    # Check if row1 contains MIN logic
                    if row1_val and 'MIN' in str(row1_val).upper():
                        cell_contribution = self._parse_min_value(row1_val, sheet_name, inputs) or 0
                    # Check if row2 has a formula that references other cells
                    elif formula_val and isinstance(formula_val, str) and formula_val.startswith('='):
                        cell_contribution = self._evaluate_cell_formula(ws, ws_formula, formula_val, row, header_row, inputs, sheet_name)
                        self._log(f"    Col {c} ({header_val}): 公式={formula_val}, 计算值={cell_contribution}")
                    elif row2_val is not None:
                        # Use row2 value directly if it's a number
                        try:
                            cell_contribution = float(row2_val)
                        except (ValueError, TypeError):
                            cell_contribution = 0
                    
                    if cell_contribution > 0:
                        self._log(f"    Col {c} ({header_val}): {cell_contribution}")
                    total_cost += cell_contribution
        else:
            # Fallback to direct value
            total_cost = ws.cell(row, e2e_cost_col).value or 0
            if not total_cost:
                total_cost = ws.cell(row + 1, e2e_cost_col).value or 0
        
        self._log(f"Total calculated cost: {total_cost}")
        
        # Get LT
        lt_row1 = ws.cell(row, e2e_lt_col).value
        lt_row2 = ws.cell(row + 1, e2e_lt_col).value
        
        if self._is_date_format(lt_row2):
            lt = lt_row2
        elif self._is_date_format(lt_row1):
            lt = lt_row1
        else:
            lt = lt_row2 if lt_row2 and str(lt_row2).strip() else lt_row1
        
        lt_str = str(lt).strip() if lt else ""
        
        breakdown, log_details = self._get_breakdown_merged(ws, ws_formula, sheet_name, row, header_row, inputs)
        return total_cost, lt_str, breakdown, log_details

    def _extract_data_from_row(self, ws, ws_formula, sheet_name, row, header_row, inputs):
        e2e_cost_col = self._get_col_by_header(ws, header_row, 'E2E Cost')
        e2e_lt_col = self._get_col_by_header(ws, header_row, 'E2E Lead Time')
        map_col = self._get_col_by_header(ws, header_row, 'MAP')
        from_col = self._get_col_by_header(ws, header_row, 'From')
        to_col = self._get_col_by_header(ws, header_row, 'To')
        
        self._log(f"E2E Cost column: {e2e_cost_col}, E2E Lead Time column: {e2e_lt_col}")
        
        # Check if this is a single-row entry
        # Method 1: Next row has a different MAP node
        current_map = ws.cell(row, map_col).value if map_col else None
        next_map = ws.cell(row + 1, map_col).value if map_col else None
        
        # Method 2: Next row has different From/To values (new record)
        current_from = ws.cell(row, from_col).value if from_col else None
        next_from = ws.cell(row + 1, from_col).value if from_col else None
        current_to = ws.cell(row, to_col).value if to_col else None
        next_to = ws.cell(row + 1, to_col).value if to_col else None
        
        is_single_row = False
        # If next row has a different MAP node, it's single-row 
        if next_map and str(next_map).strip() and current_map != next_map:
            is_single_row = True
        # If next row has different From/To (new record starts), it's single-row
        elif next_from and str(next_from).strip() and (current_from != next_from or current_to != next_to):
            is_single_row = True
        # If current row has a cost but next row's cost cell has a new MAP value, it's single-row
        elif ws.cell(row, e2e_cost_col).value and next_map and str(next_map).strip():
            is_single_row = True
        
        if is_single_row:
            self._log(f"Single-row data detected (current MAP={current_map}, next MAP={next_map})")
        
        cost = ws.cell(row, e2e_cost_col).value or 0
        self._log(f"Cost from row {row}: {cost}")
        
        # Only check row+1 for merged cells (not single-row data)
        if not cost and not is_single_row:
            cost = ws.cell(row + 1, e2e_cost_col).value or 0
            self._log(f"Cost from row {row+1}: {cost}")
        
        lt_row1 = ws.cell(row, e2e_lt_col).value
        self._log(f"LT Row {row}: '{lt_row1}'")
        
        # For single-row data, use row1 only
        if is_single_row:
            lt = lt_row1
            self._log(f"Selected LT from Row {row} (single-row data)")
        else:
            lt_row2 = ws.cell(row + 1, e2e_lt_col).value
            self._log(f"LT Row {row+1}: '{lt_row2}'")
            
            if self._is_date_format(lt_row2):
                lt = lt_row2
                self._log(f"Selected LT from Row {row+1} (date format)")
            elif self._is_date_format(lt_row1):
                lt = lt_row1
                self._log(f"Selected LT from Row {row} (date format)")
            else:
                lt = lt_row2 if lt_row2 and str(lt_row2).strip() else lt_row1
                self._log(f"Selected LT (fallback): '{lt}'")
        
        lt_str = str(lt).strip() if lt else ""
        self._log(f"Final LT: '{lt_str}'")

        breakdown, log_details = self._get_breakdown_merged(ws, ws_formula, sheet_name, row, header_row, inputs, is_single_row)
        return cost, lt_str, breakdown, log_details

    def _get_breakdown_merged(self, ws, ws_formula, sheet_name, row, header_row, inputs=None, is_single_row=False):
        base_costs = []
        variable_costs = []
        log_details = []
        
        breakdown_cols = []
        for c in range(1, ws.max_column + 1):
            idx_val = ws.cell(1, c).value
            if idx_val is not None:
                try:
                    int(idx_val)
                    breakdown_cols.append(c)
                except (ValueError, TypeError):
                    if str(idx_val).isdigit():
                        breakdown_cols.append(c)
        
        self._log(f"Breakdown columns: {breakdown_cols}")

        for c in breakdown_cols:
            title = ws.cell(header_row, c).value
            if not title: continue
            title_str = str(title).strip()
            
            fill = ws.cell(1, c).fill
            is_green = False
            if fill and hasattr(fill.start_color, 'rgb'):
                color_rgb = str(fill.start_color.rgb)
                if color_rgb in ['FF92D050', '92D050']:
                    is_green = True
            
            val1 = ws.cell(row, c).value
            
            # For single-row data, only use current row
            if is_single_row:
                val2 = val1  # Use same value for display
                # Check for formula in current row
                formula_val = ws_formula.cell(row, c).value
                if inputs and formula_val and isinstance(formula_val, str) and formula_val.startswith('='):
                    calculated_val = self._evaluate_cell_formula(ws, ws_formula, formula_val, row, header_row, inputs, sheet_name)
                    if calculated_val > 0:
                        val2 = calculated_val
            else:
                val2 = ws.cell(row + 1, c).value
                
                # Check for formula in row2 and evaluate with user inputs
                formula_val = ws_formula.cell(row + 1, c).value
                if inputs and formula_val and isinstance(formula_val, str) and formula_val.startswith('='):
                    calculated_val = self._evaluate_cell_formula(ws, ws_formula, formula_val, row, header_row, inputs, sheet_name)
                    if calculated_val > 0:
                        val2 = calculated_val
            
            # Check for MIN logic in val1
            if inputs and val1 and 'MIN' in str(val1).upper():
                calculated_val = self._parse_min_value(val1, sheet_name, inputs)
                if calculated_val > 0:
                    val2 = calculated_val
            
            val1_str = str(val1).strip() if val1 is not None else ""
            val2_str = str(val2).strip() if val2 is not None else ""
            
            # Format numbers nicely
            if val2 is not None:
                try:
                    num_val = float(val2)
                    if num_val == int(num_val):
                        val2_str = str(int(num_val))
                    else:
                        val2_str = f"{num_val:.2f}"
                except (ValueError, TypeError):
                    pass
            
            log_details.append(f"Col {c} ({title_str}): Row1='{val1_str}', Row2='{val2_str}', Green={is_green}")
            
            if val1_str or val2_str:
                item = {
                    "name": title_str, 
                    "row1": val1_str if val1_str else "-",
                    "row2": val2_str if val2_str else "-"
                }
                if is_green:
                    variable_costs.append(item)
                else:
                    base_costs.append(item)
        
        self._log(f"Base costs: {len(base_costs)} items")
        self._log(f"Variable costs: {len(variable_costs)} items")
                    
        return {"base": base_costs, "variable": variable_costs}, log_details
